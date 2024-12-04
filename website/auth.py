from flask import Blueprint, jsonify, request, flash, redirect, session, url_for, send_file, Response
from flask_login import login_user, logout_user, current_user, login_required, LoginManager
from .models import User, Organization, Report, Version_report, DirUnit, DirProduct, Sections, Ticket, Message
from . import db
from sqlalchemy import func
from sqlalchemy import desc
from sqlalchemy.orm import joinedload
from werkzeug.security import check_password_hash, generate_password_hash
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
import re
from decimal import Decimal, InvalidOperation
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import os
import pandas as pd
from tempfile import NamedTemporaryFile
import dbf
import random
import string
import zipfile
import io
import requests
from user_agents import parse
from sqlalchemy.sql import func, or_
from sqlalchemy.types import String

auth = Blueprint('auth', __name__)
login_manager = LoginManager()

@login_manager.user_loader
def load_user(id):
    return User.query.get(int(id))

@auth.route('/login', methods=['GET', 'POST'])
async def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        remember = True if request.form.get('remember') else False
        if email and password:
            user = User.query.filter(func.lower(User.email) == func.lower(email)).first()
            if user:      
                if check_password_hash(user.password, password):
                    flash('Авторизация прошла успешно', 'success')
                    login_user(user, remember=remember)
                    return redirect(url_for('views.account'))
                else:
                    flash('Неправильный пароль', 'error')    
            else:
                flash('Нет пользователя с таким email', 'error')
        else:
            flash('Введите данные для авторизации', 'error')
        return redirect(url_for('views.login'))

@auth.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Выполнен выход из аккаунта', 'success')
    return redirect(url_for('views.login'))

def get_location_info(user_agent_string):
    try:
        ip_response = requests.get("https://api64.ipify.org?format=json")
        ip_response.raise_for_status()
        ip_address = ip_response.json().get("ip")

        location_response = requests.get(f"https://ipinfo.io/{ip_address}/json")
        location_response.raise_for_status()
        location_data = location_response.json()

        user_agent = parse(user_agent_string)
        browser = user_agent.browser.family 
        os = user_agent.os.family
        
        location = location_data.get("city", "Неизвестно") + ", " + location_data.get("region", "Неизвестно") + ", " + location_data.get("country", "Неизвестно")
        ip_address_str = ip_address if ip_address else "Неизвестно"

        return ip_address_str, location, os, browser

    except Exception as e:
        print(f"Ошибка при получении данных о местоположении: {e}")
        return "Неизвестно", "Неизвестно", "Неизвестно", "Неизвестно"
    
def send_email(message_body, recipient_email, email_type, location=None, device=None, browser=None, ip_address=None):
    smtp_server = 'smtp.mail.ru'
    smtp_port = 587
    email_address = 'erespondent_online@mail.ru'
    email_password = 'fEu6vp5gSUdsUA1HpEXn'

    base_styles = """
    <style>
        body { font-family: Arial, sans-serif; background-color: #f3f3f3; margin: 0; padding: 0; }
        .email-container { max-width: 600px; margin: 20px auto; background-color: #fff; border-radius: 8px; 
                           overflow: hidden; box-shadow: 0 2px 6px rgba(0, 0, 0, 0.1); }
        .header { background-color: #f3f3f3; text-align: center; padding: 20px; }
        .header a { font-size: 32px; font-weight: bold; padding: 15px; margin: 5px 0; }
        .content { padding: 20px; color: #333; }
        .info { background-color: #f9f9f9; padding: 15px; border-left: 4px solid #028dff; margin: 20px 0; 
                border-radius: 4px; }
        .code { text-align: center; font-size: 32px; font-weight: bold; background-color: #f9f9f9; padding: 15px; 
                border: 1px solid #ddd; border-radius: 5px; margin: 20px 0; }
        .footer { background-color: #f3f3f3; padding: 10px; text-align: center; font-size: 12px; color: #777; }
        .footer a { color: #6441a5; text-decoration: none; }
    </style>
    """

    if email_type == "activation_kod":
        content = f"""
        <p>Здравствуйте</p>
        <p>Кто-то пытается войти в Erespondent-Online используя вашу электронную почту.</p>
        <div class="info">
            {f'<p><strong>Расположение:</strong> {location}</p>' if location else ''}
            {f'<p><strong>Устройство:</strong> {device}</p>' if device else ''}
            {f'<p><strong>Браузер:</strong> {browser}</p>' if browser else ''}
            {f'<p><strong>IP-адрес:</strong> {ip_address}</p>' if ip_address else ''}
        </div>
        <p>Чтобы активировать вход, введите следующий код:</p>
        <div class="code">{message_body}</div>
        <p class="note">Обратите внимание, что срок действия этого кода истекает ...</p>
        <p>Если код не применяется, запросите новый код подтверждения и попробуйте выполнить следующие действия для решения проблемы:</p>
        <ul>
            <li>Используйте режим инкогнито или другой браузер</li>
            <li>Очистите кэш вашего браузера и удалите файлы cookie</li>
            <li>Убедитесь, что браузер обновлен до последней версии</li>
        </ul>
    """
    elif email_type == "new_pass":
        content = f"""
        <p>Здравствуйте</p>
        <p>Кто-то пытается изменить пароль в Erespondent-Online используя вашу электронную почту.</p>
        <div class="info">
            {f'<p><strong>Расположение:</strong> {location}</p>' if location else ''}
            {f'<p><strong>Устройство:</strong> {device}</p>' if device else ''}
            {f'<p><strong>Браузер:</strong> {browser}</p>' if browser else ''}
            {f'<p><strong>IP-адрес:</strong> {ip_address}</p>' if ip_address else ''}
        </div>
        <p>Вот ваш новый пароль:</p>
        <div class="code">{message_body}</div>
        <p>При желании его можно сменить в профиле пользователя, для этого требуется перейти:</p>
            <ul>
                <li>Личный кабинет →</li>
                <li>Профиль →</li>
                <li>Конфиденциальность</li>
            </ul>
        """
    else:
        content = f"""
        <p>Здравствуйте</p>
        <p>Сообщение об изменении статуса отчета</p>
        <p>Статус отчета изменен на:</p>
        <div class="code">{message_body}</div>
        """

    html_template = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>{base_styles}</head>
    <body>
        <div class="email-container">
            <div class="header"><a>Erespondent-Online</a></div>
            <div class="content">
                {content}
            </div>
            <div class="footer">
                <p>Дополнительную информацию можно найти <a href="#">здесь</a>.</p>
                <p>Спасибо,<br>Служба поддержки Erespondent-Online</p>
            </div>
        </div>
    </body>
    </html>
    """

    message = MIMEMultipart("alternative")
    message['From'] = email_address
    message['To'] = recipient_email
    message['Subject'] = 'Оповещение пользователя'
    message.attach(MIMEText(html_template, 'html'))

    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(email_address, email_password)
        server.send_message(message)
        print("Письмо успешно отправлено")
    except Exception as e:
        print(f"Ошибка при отправке письма: {e}")
    finally:
        server.quit()

def send_activation_email(email):
    activation_kod = gener_password()
    session['activation_code'] = activation_kod
    send_email(activation_kod, email, 'activation_kod')

@auth.route('/sign', methods=['GET', 'POST'])
async def sign():
    if request.method == 'POST':
        email = request.form.get('email')
        password1 = request.form.get('password1')
        password2 = request.form.get('password2')
        if email and password1:
            if User.query.filter(func.lower(User.email) == func.lower(email)).first():
                flash('Пользователь с таким email уже существует', 'error')
            elif not re.match(r'[\w\.-]+@[\w\.-]+', email):
                flash('Некорректный адрес электронной почты', 'error') 
            elif password1 != password2:
                flash('Ошибка в подтверждении пароля', 'error') 
            else:      
                session['temp_user'] = {
                    'email': email,
                    'password': generate_password_hash(password1)
                }
                send_activation_email(email) 
                flash('Регистрация прошла успешно! Проверьте свою почту для активации аккаунта.', 'success')
                return redirect(url_for('views.kod'))
        else:
            flash('Введите данные для регистрации', 'error')    
    return redirect(url_for('views.sign'))

@auth.route('/kod', methods=['GET', 'POST'])
def kod():
    if request.method == 'POST':
        input_code = ''.join([
            request.form.get(f'activation_code_{i}', '') for i in range(5)
        ])
        if input_code == session.get('activation_code'):
            new_user = User(
                email=session['temp_user']['email'],
                password=session['temp_user']['password']
            )
            db.session.add(new_user)
            db.session.commit()
            session.pop('temp_user', None)
            session.pop('activation_code', None)
            flash('Аккаунт успешно активирован!', 'success')
            return redirect(url_for('views.login'))
        else:
            flash('Некорректный код активации', 'error')
    return redirect(url_for('views.kod'))

@auth.route('/resend_code', methods=['POST'])
def resend_code():
    email = session.get('temp_user', {}).get('email')
    if email:
        new_activation_code = gener_password()
        session['activation_code'] = new_activation_code
        send_email(new_activation_code, email, 'activation_kod')
        return jsonify({'status': 'success', 'message': 'Код активации отправлен повторно!'})
    else:
        return jsonify({'status': 'error', 'message': 'Не удалось отправить код повторно.'}), 400

@auth.route('/add_personal_parametrs', methods=['GET', 'POST'])
async def add_personal_parametrs():
    if request.method == 'POST':
        name = request.form.get('name_common')
        second_name = request.form.get('second_name_common')
        patronymic = request.form.get('patronymic_common')
        telephone = request.form.get('telephone_common')   
        full_name = request.form.get('full_name_common')
        organiz_full_name = Organization.query.filter_by(full_name = full_name).first()
        if not name or not telephone or not full_name or not name or not second_name:
            flash('Заполните все обязательные строки', 'error')
            return redirect(url_for('views.profile_common'))
        else:
            name = name.replace(' ', '')
            second_name = second_name.replace(' ', '')
            patronymic = patronymic.replace(' ', '')
            fio = f"{second_name} {name} {patronymic}"
            current_user.fio = fio
            db.session.commit()
            existing_telephone = User.query.filter(User.id != current_user.id, User.telephone == telephone).first()
            if existing_telephone:
                flash('Пользователь с таким номером телефона уже существует.', 'error')
            else: 
                current_user.telephone = telephone
                db.session.commit()
                existing_userOrg = User.query.filter_by(organization_id=organiz_full_name.id).first()
                if existing_userOrg:
                    flash('Аккаунт с такой организацией уже существует.', 'error')
                else: 
                    current_user.organization_id = organiz_full_name.id
                    db.session.commit()
                    flash('Данные успешно обновлены.', 'success')
                    return redirect(url_for('views.profile_common'))  
        return redirect(url_for('views.profile_common'))

def update_user_activity():
    if current_user.is_authenticated:
        current_user.update_activity()

@auth.before_request
def before_request():
    update_user_activity()

@auth.route('/update_activity', methods=['POST'])
def update_activity():
    if current_user.is_authenticated:
        current_user.update_activity()
    return '', 204

@auth.route('/profile/password', methods=['GET', 'POST'])
async def profile_password():
    if request.method == 'POST':
        old_password = request.form.get('old_password')
        new_password = request.form.get('new_password')
        conf_new_password = request.form.get('conf_new_password')
        user = User.query.filter_by(email=current_user.email).first()
        if old_password and new_password and conf_new_password:
            if not check_password_hash(current_user.password, old_password):
                flash('Не правильный старый пароль', 'error')
                return redirect(url_for('views.my_profile'))
            elif new_password != conf_new_password:
                flash('При подтверждении пароля произошла ошибка', 'error')
                return redirect(url_for('views.my_profile'))
            else:
                user.password = generate_password_hash(conf_new_password)
                db.session.commit()
                flash('Пароль был изменен', 'success')
                send_email(conf_new_password, current_user.email, 'new_pass')
                return redirect(url_for('views.login'))
        else:
            flash('Введите пароль ', 'error')
    return redirect(url_for('views.profile_password'))

def gener_password():
    length=5
    characters = string.digits
    # characters = string.ascii_letters + string.digits
    password = ''.join(random.choice(characters) for _ in range(length))
    return password

@auth.route('/relod_password', methods=['POST'])
async def relod_password():
    email = request.form.get('email_relod')
    if not email:
        flash('Заполните строку с E-mail для отправки нового пароля', 'error')
        return redirect(url_for('views.login'))
    user = User.query.filter(func.lower(User.email) == func.lower(email)).first()
    if user:
        new_password = gener_password()
        hashed_password = generate_password_hash(new_password)

        user_agent_string = request.headers.get('User-Agent')
        ip_address, location, os, browser = get_location_info(user_agent_string)
        
        send_email(new_password, email, 'new_pass', location=location, device=os, browser=browser, ip_address=ip_address)
        flash('Новый пароль был отправлен вам на email', 'success')
        user.password = hashed_password
        db.session.commit()
        return redirect(url_for('views.login'))
    else:
        flash('Пользователя с таким email не существует', 'error')
        return redirect(url_for('views.login'))

@auth.route('/create_new_report', methods=['POST'])
async def create_new_report():
    if request.method == 'POST':    
        organization_name = request.form.get('modal_organization_name')
        okpo = request.form.get('modal_organization_okpo')
        year = request.form.get('modal_report_year')
        quarter = request.form.get('modal_report_quarter')
        organization = Organization.query.filter_by(okpo = okpo, full_name=organization_name).first()
        proverka_report = Report.query.filter_by(organization_name=organization_name, year = year, quarter=quarter).first()
        if not proverka_report:
            new_report = Report(
                okpo=organization.okpo,
                organization_name= organization.full_name,
                year=year,
                quarter=quarter,
                user_id = current_user.id
            )
            db.session.add(new_report)
            db.session.commit()
            new_version_report = Version_report(
                begin_time = datetime.now(), 
                status = "Заполнение",
                fio = current_user.fio,
                telephone = current_user.telephone,
                email = current_user.email,
                report=new_report
            )
            db.session.add(new_version_report)
            db.session.commit() 

            sections = Sections.query.filter_by(id_version=new_version_report.id).all()
            if not sections:
                id = new_version_report.id
                sections_data = [
                    (id, 288, 9100, 1, '', 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, ''),
                    (id, 285, 9010, 1, '', 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, ''),
                    (id, 282, 9001, 1, '', 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, ''),
      
                    (id, 290, 9100, 2, '', 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, ''),
                    (id, 287, 9010, 2, '', 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, ''),
                    (id, 284, 9001, 2, '', 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, ''),
        
                    (id, 289, 9100, 3, '', 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, ''),
                    (id, 286, 9010, 3, '', 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, ''),
                    (id, 283, 9001, 3, '', 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, ''),
                ]
                for data in sections_data:
                    section = Sections(
                        id_version=data[0],
                        id_product=data[1],
                        code_product=data[2],
                        section_number=data[3],
                        Oked=data[4],
                        produced=data[5],
                        Consumed_Quota=data[6],
                        Consumed_Fact=data[7],
                        Consumed_Total_Quota=data[8],
                        Consumed_Total_Fact=data[9],
                        total_differents=data[10],
                        note=data[11]
                    )
                    db.session.add(section)
                db.session.commit()
            flash('Добавлен новый отчет', 'success')
        else:
            flash(f'Отчет {year} года {quarter} квартала уже существует', 'error')
    return redirect(url_for('views.report_area'))
    
def last_quarter():
    current_mounth = datetime.now().strftime("%m")
    if (current_mounth == '01' or current_mounth == '02' or current_mounth == '03'):
        last_quarter_value = 4
    elif (current_mounth == '04' or current_mounth == '05' or current_mounth == '06'):
        last_quarter_value = 1
    elif (current_mounth == '07' or current_mounth == '08' or current_mounth == '09'):
        last_quarter_value = 2
    else:
        last_quarter_value = 3
    return last_quarter_value
  
def year_fourMounth_ago():
    current_date = datetime.now()
    months_to_subtract = 4
    new_date = current_date - timedelta(days=months_to_subtract * 30)
    year_4_months_ago = new_date.year
    return year_4_months_ago

@auth.route('/update_report', methods=['POST'])
async def update_report():
    if request.method == 'POST':
        id = request.form.get('modal_report_id')
        okpo = request.form.get('modal_report_okpo')
        year = request.form.get('modal_change_report_year')
        quarter = request.form.get('modal_change_report_quarter')   
        current_report = Report.query.filter_by(id=id).first()
        versions = Version_report.query.filter_by(report_id=id).all()
        organization_okpo = Organization.query.filter_by(okpo=okpo).first()

        if current_report:
            existing_report = Report.query.filter_by(
                okpo=okpo,
                year=year,
                quarter=quarter
            ).first()

            if existing_report and existing_report.id != id:
                flash('Отчет с таким годом и кварталом уже существует', 'error')
                return redirect(url_for('views.report_area'))

            sent_version_exists = any(version.status == 'Отправлен' for version in versions)
            if sent_version_exists:
                flash('После отправки изменение отчета недоступно', 'error')
                return redirect(url_for('views.report_area'))
            
            confirmed_version_exists = any(version.status == 'Одобрен' for version in versions)
            if confirmed_version_exists:
                flash('Данный отчет не подлежит редактированию', 'error')
                return redirect(url_for('views.report_area'))

            current_report.okpo = okpo
            current_report.year = year
            current_report.quarter = quarter
            current_report.organization_name = organization_okpo.full_name
            db.session.commit()
            flash('Параметры обновлены', 'success')
        else:
            flash('Отчет не найден', 'error')

        return redirect(url_for('views.report_area'))
    
@auth.route('/сopy_report', methods=['POST'])
async def сopy_report():
    if request.method == 'POST':
        coppy_report_id = request.form.get('coppy_report_id')
        current_report_version = Version_report.query.filter_by(report_id = coppy_report_id).first()
        current_sections = Sections.query.filter_by(id_version=current_report_version.id).all()
        new_organization_name = request.form.get('coppy_organization_name')
        new_organization_okpo = request.form.get('coppy_organization_okpo')
        new_report_year = request.form.get('coppy_report_year')
        new_report_quarter = request.form.get('coppy_report_quarter')
        proverka_report = Report.query.filter_by(organization_name=new_organization_name, 
                                                 year = new_report_year, 
                                                 quarter=new_report_quarter, 
                                                 okpo = new_organization_okpo).first()
        if not proverka_report:
            new_report = Report(
                okpo=new_organization_okpo,
                organization_name= new_organization_name,
                year=new_report_year,
                quarter=new_report_quarter,
                user_id = current_user.id
            )
            db.session.add(new_report)
            db.session.commit()
            new_version = Version_report(
                status = "Заполнение",
                fio = current_user.fio,
                telephone = current_user.telephone,
                email = current_user.email,
                report=new_report
            )
            db.session.add(new_version)
            db.session.commit()
            for section in current_sections:
                new_section = Sections(
                    id_version=new_version.id,
                    id_product=section.id_product,
                    code_product=section.code_product,
                    section_number=section.section_number,
                    Oked=section.Oked,
                    produced=section.produced,
                    Consumed_Quota=section.Consumed_Quota,
                    Consumed_Fact=section.Consumed_Fact,
                    Consumed_Total_Quota=section.Consumed_Total_Quota,
                    Consumed_Total_Fact=section.Consumed_Total_Fact,
                    total_differents=section.total_differents,
                    note=section.note
                )
                db.session.add(new_section)
            db.session.commit()
        else:
            flash('Отчет с таким годом и квараталом уже существует, копирование eror','error')
        return redirect(url_for('views.report_area'))

@auth.route('/delete_report/<report_id>', methods=['POST'])
async def delete_report(report_id):
    if request.method == 'POST':
        current_report = Report.query.filter_by(id = report_id).first()
        versions = Version_report.query.filter_by(report_id = report_id).all()
        tickets = Ticket.query.filter_by(version_report_id = report_id).all()  
        if current_report:   
            sent_version_exists = any(version.status == 'Отправлен' for version in versions)
            if sent_version_exists:
                flash('Отправленный отчет не подлежит удалению', 'error')
                return redirect(url_for('views.report_area'))  
            confirmed_version_exists = any(version.status == 'Одобрен' for version in versions)
            if confirmed_version_exists:
                flash('Данный отчет не подлежит удалению', 'error')
                return redirect(url_for('views.report_area'))
            for ticket in tickets:
                db.session.delete(ticket)        
            for version in versions:
                sections = Sections.query.filter_by(id_version = version.id).all()
                for section in sections:
                    db.session.delete(section)
                db.session.delete(version)
            db.session.delete(current_report)
            db.session.commit()
            flash('Отчет удален', 'success')
        return redirect(url_for('views.report_area'))
    
@auth.route('/create_new_report_version/<int:id>', methods=['POST'])
async def create_new_report_version(id):
    if request.method == 'POST':
        new_version_report = Version_report(
            fio = current_user.fio,
            telephone = current_user.telephone,
            email = current_user.email,
            report_id = id 
        )
        db.session.add(new_version_report)
        db.session.commit()
        flash('Добавлена новая версия', category='success')
    return redirect(url_for('views.report_area'))

@auth.route('/delete_version/<int:id>', methods=['POST'])
async def delete_version(id):
    if request.method == 'POST':
        current_version = Version_report.query.filter_by(id=id).first()
        if current_version:
            if current_version.status != 'Отправлен':
                versions_currentReport = Version_report.query.filter_by(report_id=current_version.report_id).all()
                if len(versions_currentReport) > 1:
                    sections = Sections.query.filter_by(id_version=id).all()
                    tickets = Ticket.query.filter_by(version_report_id=id).all()
                    for section in sections:
                        db.session.delete(section)
                    for ticket in tickets:
                        db.session.delete(ticket)
                    db.session.delete(current_version)
                    db.session.commit()
                    flash('Версия удалена', category='success')
                else:
                    flash('Единственная версия не может быть удалена', category='error')
            else:
                flash('Отправленная версия не может быть удалена', category='error')
        else:
            flash('Версия не найдена', category='error')
    return redirect(url_for('views.report_area'))

@auth.route('/add_section_param', methods=['POST'])
async def add_section_param():
    if request.method == 'POST':
        current_version_id = request.form.get('current_version')
        name = request.form.get('name_of_product')
        id_product = request.form.get('add_id_product')
        oked = request.form.get('oked_add')
        produced = request.form.get('produced_add')
        Consumed_Quota = request.form.get('Consumed_Quota_add')
        Consumed_Fact = request.form.get('Consumed_Fact_add')
        Consumed_Total_Quota = request.form.get('Consumed_Total_Quota_add')
        Consumed_Total_Fact = request.form.get('Consumed_Total_Fact_add')
        note = request.form.get('note_add')
        section_number = request.form.get('section_number')
        
        produced = Decimal(produced) if produced else Decimal(0)
        Consumed_Quota = Decimal(Consumed_Quota) if Consumed_Quota else Decimal(0)
        Consumed_Fact = Decimal(Consumed_Fact) if Consumed_Fact else Decimal(0)
        Consumed_Total_Quota = Decimal(Consumed_Total_Quota) if Consumed_Total_Quota else Decimal(0)
        Consumed_Total_Fact = Decimal(Consumed_Total_Fact) if Consumed_Total_Fact else Decimal(0)

        current_version = Version_report.query.filter_by(id=current_version_id).first()
        current_product = DirProduct.query.filter_by(NameProduct=name, IdProduct=id_product).first()
        product_unit = DirUnit.query.filter_by(IdUnit=current_product.IdUnit).first() if current_product else None
        if current_version:
            if current_version.status != 'Отправлен' and current_version.status != 'Одобрен':
                if current_product:
                    proverka_section = Sections.query.filter_by(id_version=current_version_id, section_number=section_number, id_product=current_product.IdProduct).first()
                    if proverka_section:
                        flash('Такой вид продукции уже существует', 'error')
                        if section_number == '1':
                            return redirect(url_for('views.report_section', report_type='fuel', id=current_version_id))
                        elif section_number == '2':
                            return redirect(url_for('views.report_section', report_type='heat', id=current_version_id))
                        elif section_number == '3':
                            return redirect(url_for('views.report_section', report_type='electro', id=current_version_id))
                    else:
                        new_section = Sections(
                            id_version=current_version_id,
                            id_product=current_product.IdProduct,
                            code_product=current_product.CodeProduct,
                            section_number=section_number,
                            Oked=oked,
                            produced=produced,
                            Consumed_Quota=Consumed_Quota,
                            Consumed_Fact=Consumed_Fact,
                            Consumed_Total_Quota=Consumed_Total_Quota,
                            Consumed_Total_Fact=Consumed_Total_Fact,
                            total_differents=None,
                            note=note
                        )
                        db.session.add(new_section)
                        db.session.commit()
                        current_section = Sections.query.filter_by(id=new_section.id, section_number=section_number).first()

                        if current_section.code_product == "7000" or current_section.code_product == "0020"  or current_section.code_product == "0021" or current_section.code_product == "0024" or current_section.code_product == "0025" or current_section.code_product == "0026" or current_section.code_product == "0027" or current_section.code_product == "0030" or current_section.code_product == "0031":
                            current_section.total_differents = current_section.Consumed_Total_Fact - current_section.Consumed_Total_Quota
                            db.session.commit()
                        else:
                            try:
                                if current_section.produced != 0:
                                    if product_unit and (product_unit.NameUnit == '%' or product_unit.NameUnit == '% (включая покупную)'):
                                        current_section.Consumed_Fact = round((current_section.Consumed_Total_Fact / current_section.produced) * 100, 2)
                                    else:
                                        current_section.Consumed_Fact = round((current_section.Consumed_Total_Fact / current_section.produced) * 1000, 2)
                                else:
                                    current_section.Consumed_Fact = 0

                                if current_section.Consumed_Quota != 0:
                                    if product_unit and (product_unit.NameUnit == '%' or product_unit.NameUnit == '% (включая покупную)'):
                                        current_section.Consumed_Total_Quota = round((current_section.produced / current_section.Consumed_Quota) * 100, 2)
                                    else:
                                        current_section.Consumed_Total_Quota = round((current_section.produced / current_section.Consumed_Quota) * 1000, 2)
                                else:
                                    current_section.Consumed_Total_Quota = 0
                                current_section.total_differents = current_section.Consumed_Total_Fact - current_section.Consumed_Total_Quota
                                db.session.commit()
                            except InvalidOperation as e:
                                flash(f"Ошибка при вычислениях: {e}")

                        specific_codes = ['9001', '9010', '9100']
                        section9001 = Sections.query.filter_by(id_version=current_version_id, section_number=section_number, code_product='9001').first()
                        aggregated_values = db.session.query(
                            func.sum(Sections.Consumed_Total_Quota),
                            func.sum(Sections.Consumed_Total_Fact),
                            func.sum(Sections.total_differents)
                        ).filter(
                            Sections.id_version == current_version_id,
                            Sections.section_number == section_number,
                            ~Sections.code_product.in_(specific_codes)
                        ).first()

                        if section9001 and aggregated_values:
                            section9001.Consumed_Total_Quota, section9001.Consumed_Total_Fact, section9001.total_differents = aggregated_values
                            db.session.commit()

                        section9010 = Sections.query.filter_by(id_version=current_version_id, section_number=section_number, code_product='9010').first()
                        section9100 = Sections.query.filter_by(id_version=current_version_id, section_number=section_number, code_product='9100').first()

                        if section9100 and section9001 and section9010:
                            section9100.Consumed_Total_Quota = section9001.Consumed_Total_Quota + section9010.Consumed_Total_Quota
                            section9100.Consumed_Total_Fact = section9001.Consumed_Total_Fact + section9010.Consumed_Total_Fact
                            section9100.total_differents = section9001.total_differents + section9010.total_differents
                            db.session.commit()
                    
                        if current_version:
                            current_version.change_time = datetime.now()
                            db.session.commit()
                    flash('Продукция была добавлена', 'success')
                    current_version.status = "Заполнение"
                    current_version.sent_time = None
                    db.session.commit()
                else:
                    flash('Выбирите вид продукции из выпадающего списка', 'error')
            else:
                flash('Редактирование отправленного/одобренного отчета недоступно', 'error')
        else:
            flash('Версия не найдена', 'error') 
    if section_number == '1':
        return redirect(url_for('views.report_section', report_type='fuel', id=current_version_id))
    elif section_number == '2':
        return redirect(url_for('views.report_section', report_type='heat', id=current_version_id))
    elif section_number == '3':
        return redirect(url_for('views.report_section', report_type='electro', id=current_version_id))

@auth.route('/change_section', methods=['POST'])
async def change_section():
    if request.method == 'POST':
        id_version = request.form.get('current_version')
        id_fuel = request.form.get('id')
        produced = request.form.get('produced_change')
        Consumed_Quota = request.form.get('Consumed_Quota_change')     
        Consumed_Fact = request.form.get('Consumed_Fact_change')
        Consumed_Total_Quota = request.form.get('Consumed_Total_Quota_change')
        Consumed_Total_Fact = request.form.get('Consumed_Total_Fact_change')
        note = request.form.get('note_change')

        produced = Decimal(produced) if produced else Decimal(0)
        Consumed_Quota = Decimal(Consumed_Quota) if Consumed_Quota else Decimal(0)
        Consumed_Fact = Decimal(Consumed_Fact) if Consumed_Fact else Decimal(0)
        Consumed_Total_Quota = Decimal(Consumed_Total_Quota) if Consumed_Total_Quota else Decimal(0)
        Consumed_Total_Fact = Decimal(Consumed_Total_Fact) if Consumed_Total_Fact else Decimal(0)

        current_version = Version_report.query.filter_by(id=id_version).first() 
        current_section = Sections.query.filter_by(id=id_fuel).first()
        
        if current_version:
            if current_version.status != 'Отправлен' and current_version.status != 'Одобрен':
                if current_section:
                    if current_section.code_product == "7000" or current_section.code_product == "0020"  or current_section.code_product == "0021" or current_section.code_product == "0024" or current_section.code_product == "0025" or current_section.code_product == "0026" or current_section.code_product == "0027" or current_section.code_product == "0030" or current_section.code_product == "0031":
                        current_section.Consumed_Total_Quota = Consumed_Total_Quota
                        current_section.Consumed_Total_Fact = Consumed_Total_Fact
                        current_section.note = note
                        db.session.commit()
                        current_section.total_differents = current_section.Consumed_Total_Fact - current_section.Consumed_Total_Quota
                        db.session.commit()
                    else:
                        current_section.produced = produced
                        current_section.Consumed_Quota = Consumed_Quota     
                        current_section.Consumed_Total_Fact = Consumed_Total_Fact
                        current_section.note = note 
                        db.session.commit()
                        if current_section.produced != 0:
                            current_section.Consumed_Fact = round((current_section.Consumed_Total_Fact / current_section.produced) * 1000, 2)
                        else:
                            current_section.Consumed_Fact = 0.00
                        if current_section.Consumed_Quota != 0:
                            current_section.Consumed_Total_Quota = round((current_section.produced / current_section.Consumed_Quota) * 1000, 2)
                        else:
                            current_section.Consumed_Total_Quota = 0.00
                        db.session.commit()
                        current_section.total_differents = current_section.Consumed_Total_Fact - current_section.Consumed_Total_Quota
                        db.session.commit()
                    if current_version:
                        current_version.change_time = datetime.now()
                        db.session.commit()

                    specific_codes = ['9001', '9010', '9100']
                    section9001 = Sections.query.filter_by(id_version=id_version, section_number=current_section.section_number, code_product=9001).first()
                    aggregated_values = db.session.query(
                        func.sum(Sections.Consumed_Total_Quota),
                        func.sum(Sections.Consumed_Total_Fact),
                        func.sum(Sections.total_differents)
                    ).filter(
                        Sections.id_version == id_version,
                        Sections.section_number == current_section.section_number,
                        ~Sections.code_product.in_(specific_codes)
                    ).first()

                    if aggregated_values:
                        section9001.Consumed_Total_Quota = aggregated_values[0] or 0
                        section9001.Consumed_Total_Fact = aggregated_values[1] or 0
                        section9001.total_differents = aggregated_values[2] or 0
                        db.session.commit()

                    section9010 = Sections.query.filter_by(id_version=id_version, section_number=current_section.section_number, code_product=9010).first()
                    section9100 = Sections.query.filter_by(id_version=id_version, section_number=current_section.section_number, code_product=9100).first()
                    if section9100 and section9001 and section9010:
                        section9100.Consumed_Total_Quota = (section9001.Consumed_Total_Quota or 0) + (section9010.Consumed_Total_Quota or 0)
                        section9100.Consumed_Total_Fact = (section9001.Consumed_Total_Fact or 0) + (section9010.Consumed_Total_Fact or 0)
                        section9100.total_differents = (section9001.total_differents or 0) + (section9010.total_differents or 0)
                        db.session.commit()
                    current_version.status = "Заполнение"
                    current_version.sent_time = None
                    db.session.commit()
                    flash('Параметры обновлены', 'success')
                else:
                    flash('Ошибка при обновлении', 'error')   
            else:
                flash('Редактирование отправленного/одобренного отчета недоступно', 'error')
        else:
            flash('Версия не найдена', 'error')
        if(current_section.section_number == 1):
            return redirect(url_for('views.report_section', report_type='fuel', id=id_version))
        elif(current_section.section_number == 2):
            return redirect(url_for('views.report_section', report_type='heat', id=id_version))
        elif(current_section.section_number == 3):
            return redirect(url_for('views.report_section', report_type='electro', id=id_version))

@auth.route('/remove_section/<id>', methods=['POST'])
async def remove_section(id):
    if request.method == 'POST':
        delete_section = Sections.query.filter_by(id=id).first()
        id_version = delete_section.id_version
        section_numberDELsection = delete_section.section_number
        current_version = Version_report.query.filter_by(id=id_version).first() 
        if current_version:
            if current_version.status != 'Отправлен' and current_version.status != 'Одобрен':
                if delete_section:
                    section9001 = Sections.query.filter_by(id_version=id_version, section_number = section_numberDELsection, code_product = 9001).first()
                    section9001.Consumed_Total_Quota -= delete_section.Consumed_Total_Quota
                    section9001.Consumed_Total_Fact -= delete_section.Consumed_Total_Fact
                    section9001.total_differents -= delete_section.total_differents

                    section9100 = Sections.query.filter_by(id_version=id_version, section_number = section_numberDELsection, code_product = 9100).first()
                    section9100.Consumed_Total_Quota -= delete_section.Consumed_Total_Quota
                    section9100.Consumed_Total_Fact -= delete_section.Consumed_Total_Fact
                    section9100.total_differents -= delete_section.total_differents

                    db.session.delete(delete_section)
                    db.session.commit()
                    flash('Продукция была удалена', 'success')
                    if current_version:
                        current_version.change_time = datetime.now()
                        db.session.commit()
                    current_version.status = "Заполнение"
                    current_version.sent_time = None
                    db.session.commit()
                else: 
                    flash('Ошибка при удалении', 'error')
            else:
                flash('Редактирование отправленного/одобренного отчета недоступно', 'error')  
        else:
            flash('Версия не найдена', 'error')

        if(section_numberDELsection == 1):
            return redirect(url_for('views.report_section', report_type='fuel', id=id_version))
        elif(section_numberDELsection == 2):
            return redirect(url_for('views.report_section', report_type='heat', id=id_version))
        elif(section_numberDELsection == 3):
            return redirect(url_for('views.report_section', report_type='electro', id=id_version))
        
@auth.route('/control_version/<id>', methods=['POST'])
async def control_version(id):
    if request.method == 'POST':
        current_version = Version_report.query.filter_by(id=id).first()
        id_version = current_version.id
        
        sections = {
            'fuel': Sections.query.filter_by(id_version=id, section_number=1, code_product=9010).first(),
            'heat': Sections.query.filter_by(id_version=id, section_number=2, code_product=9010).first(),
            'electro': Sections.query.filter_by(id_version=id, section_number=3, code_product=9010).first(),
        }

        if current_version.status == 'Заполнение': 
            for key, section in sections.items():
                if section is None or not section.note:
                    flash('Примечание с кодом строки "9010" обязательно для заполнения', 'error')
                    return redirect(url_for('views.report_section', report_type=key, id=id_version))
            
            current_version.status = 'Контроль пройден'
            db.session.commit()
            flash('Контроль пройден', 'successful')
        else:
            flash('Контроль уже был пройден', 'error')
        return redirect(url_for('views.report_area'))
    
@auth.route('/agreed_version/<id>', methods=['POST'])
async def agreed_version(id):
    if request.method == 'POST':
        current_version = Version_report.query.filter_by(id=id).first()
        if current_version.status == 'Контроль пройден':     
            current_version.status = 'Согласовано'
            db.session.commit()
            flash('Отчет согласован', 'successful')
        elif current_version.status == 'Согласовано': 
            flash('Отчет уже согласован', 'succeful')
        else:
            flash('Необходимо пройти контроль', 'error')
        return redirect(url_for('views.report_area'))
    
@auth.route('/sent_version/<id>', methods=['POST'])
async def sent_version(id):
    if request.method == 'POST':
        current_version = Version_report.query.filter_by(id=id).first()
        if current_version.status == 'Согласовано':
            current_version.status = 'Отправлен'
            if current_version.sent_time == None:
                current_version.sent_time = datetime.now()
            db.session.commit()
            flash('Отчет отправлен', 'successful')

            user_message = Message(
                text = f"Отчет: №{current_version.id} был отправлен на проверку",
                user = current_user
            )

            db.session.add(user_message)
            db.session.commit()
        elif current_version.status == 'Отправлен':
            flash('Отчет уже был отправлен на проверку', 'error')
        else:
            flash('Необходимо согласовать', 'error')
        return redirect(url_for('views.report_area'))

@auth.route('/change_category_report', methods=['POST'])
async def change_category_report():
    action = request.form.get('action')
    report_id = request.form.get('reportId')

    status_itog = None
    if request.method == 'POST':
        current_version = Version_report.query.filter_by(id=report_id).first()
        if current_version is not None: 
            user = User.query.filter_by(email=current_version.email).first()
            if not current_version.hasNot and action != 'to_download':
                flash('Необходимо уточнить о каких ошибках идет речь', 'error')
                return redirect(url_for('views.audit_report', id=current_version.id)) 
            
            if action == 'not_viewed':
                status_itog = 'Отправлен'
            elif action == 'remarks':
                status_itog = 'Есть замечания'
                user_message = Message(
                    text=f"При проверке отчета №{current_version.id} были найдены ошибки. Исправьте ошибки и отправьте повторно.",
                    user=user
                )
                db.session.add(user_message)
            elif action == 'to_download':
                status_itog = 'Одобрен'
                user_message = Message(
                    text=f"Отчет №{current_version.id} был передан в следующую стадию проверки.",
                    user=user
                )
                db.session.add(user_message)
                ticket_message = Ticket(
                    note="Ошибок нет, отчет передан в следующую стадию проверки.",
                    luck=True,
                    version_report_id=current_version.id
                )
                db.session.add(ticket_message)
            elif action == 'to_delete':
                status_itog = 'Готов к удалению'
                user_message = Message(
                    text=f"Отчет №{current_version.id} не подлежит рассмотрению.",
                    user=user
                )
                db.session.add(user_message)

            else:
                flash('Неизвестное действие', 'error')
                return redirect(request.referrer) 
            current_version.hasNot = False
            current_version.status = status_itog
            db.session.commit()

            
            send_email(status_itog, user.email, 'change_status')
            flash(f'Статус отчета №{current_version.id} был изменен', 'success')
            return redirect(request.referrer) 
        else:
            flash('Отчет не найден', 'error')
            return "Version not found", 404
        
@auth.route('/rollbackreport/<id>', methods=['POST'])
async def rollbackreport(id):
    if request.method == 'POST':
        current_version = Version_report.query.filter_by(id=id).first()
        current_report = Report.query.filter_by(id=id).first()
        current_user = current_report.user_id

        current_version.status = "Отправлен"
        db.session.commit()
        flash(f'Статус отчета был изменен на непросмотренный', 'success')
    return redirect(request.referrer) 

@auth.route('/send_comment', methods=['POST'])
async def send_comment():
    if request.method == 'POST':
        version_id = request.form.get('version_id')
        resp_email = request.form.get('resp_email')
        text = request.form.get('text')
        cleaned_text = text
        # cleaned_text = ' '.join(text.split())
        current_version = Version_report.query.filter_by(id=version_id).first()

        if current_version:
            new_comment = Ticket(
                # begin_time = 
                # luck = 
                note = cleaned_text,
                version_report_id = current_version.id
            )
            db.session.add(new_comment)
            current_version.hasNot = True
            db.session.commit()
            flash('Комментарий создан', 'success')
        else:
            flash('Отчет не найден', 'error')
        return redirect(url_for('views.audit_report', id = version_id))

@auth.route('/export_table', methods=['POST'])
async def export_table():
    if request.method == 'POST':
        version_id = int(request.form.get('version_id'))

        sections1 = Sections.query.filter_by(id_version=version_id, section_number=1).order_by(desc(Sections.id)).all()
        sections2 = Sections.query.filter_by(id_version=version_id, section_number=2).order_by(desc(Sections.id)).all()
        sections3 = Sections.query.filter_by(id_version=version_id, section_number=3).order_by(desc(Sections.id)).all()

        unit_headers = {
            1: ('кг у.т.', 'т у.т.'),
            2: ('Мкал', 'Гкал'),
            3: ('кВтч', 'тыс.кВтч')
        }

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        def add_sheet(ws, sections, title, unit_header_one_text, unit_header_all_text):
            ws.title = title

            merged_cells = {
                'A1:A2': "Наименование вида продукции (работ услуг)",
                'B1:B2': "Код строки",
                'C1:C2': "Код по ОКЭД",
                'D1:D2': "Единица измерения",
                'E1:E2': "Произведено продукции (работ, услуг) за отчетный период",
                'F1:G1': f"Израсходовано на единицу продукции (работы, услуги) за отчетный период, {unit_header_one_text}",
                'H1:J1': f"Израсходовано на всю произведенную продукцию (работу, услугу) за отчетный период, {unit_header_all_text}",
                'K1:K2': "Примечание",
                'J1': "Экономия(-), перерасход(+)",
                'F2': "по утвержденной норме (предельному уровню)",
                'G2': "фактически",
                'H2': "по утвержденной норме (предельному уровню)",
                'I2': "фактически",
                'J2': "Экономия(-), перерасход(+)",
                'A3': "A",
                'B3': "Б",
                'C3': "В",
                'D3': "Г",
                'E3': "1",
                'F3': "2",
                'G3': "3",
                'H3': "4",
                'I3': "5",
                'J3': "6",
                'K3': "7"
            }

            font = Font(name='Times New Roman', size=12)
            alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )

            for cell_range, text in merged_cells.items():
                top_left_cell = cell_range.split(':')[0]
                ws[top_left_cell] = text
                ws[top_left_cell].font = font
                ws[top_left_cell].alignment = alignment
                ws[top_left_cell].border = border

            for cell_range in merged_cells.keys():
                ws.merge_cells(cell_range)

            column_widths = {
                'A': 30,
                'B': 15,
                'C': 15,
                'D': 20,
                'E': 50,
                'F': 40,
                'G': 40,
                'H': 40,
                'I': 40,
                'J': 30,
                'K': 20
            }
            row_heights = {
                1: 30,
                2: 30,
                3: 25
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            for row, height in row_heights.items():
                ws.row_dimensions[row].height = height

            row_index = 4
            for section in sections:
                ws[f'A{row_index}'] = section.product.NameProduct
                ws[f'B{row_index}'] = section.code_product
                ws[f'C{row_index}'] = section.Oked
                ws[f'D{row_index}'] = section.product.unit.NameUnit
                ws[f'E{row_index}'] = section.produced
                ws[f'F{row_index}'] = section.Consumed_Quota
                ws[f'G{row_index}'] = section.Consumed_Fact
                ws[f'H{row_index}'] = section.Consumed_Total_Quota
                ws[f'I{row_index}'] = section.Consumed_Total_Fact
                ws[f'J{row_index}'] = section.total_differents
                ws[f'K{row_index}'] = section.note

                for col in column_widths.keys():
                    cell = ws[f'{col}{row_index}']
                    cell.border = border

                row_index += 1

        unit_header_one_text1, unit_header_all_text1 = unit_headers.get(1, ('', ''))
        unit_header_one_text2, unit_header_all_text2 = unit_headers.get(2, ('', ''))
        unit_header_one_text3, unit_header_all_text3 = unit_headers.get(3, ('', ''))

        add_sheet(wb.create_sheet(), sections1, "Топливо", unit_header_one_text1, unit_header_all_text1)
        add_sheet(wb.create_sheet(), sections2, "Тепло", unit_header_one_text2, unit_header_all_text2)
        add_sheet(wb.create_sheet(), sections3, "Электричество", unit_header_one_text3, unit_header_all_text3)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True, download_name='report.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@auth.route('/export_version/<id>', methods=['POST'])
async def export_version(id):
    if request.method == 'POST':
        version_id = id

        sections1 = Sections.query.filter_by(id_version=version_id, section_number=1).order_by(desc(Sections.id)).all()
        sections2 = Sections.query.filter_by(id_version=version_id, section_number=2).order_by(desc(Sections.id)).all()
        sections3 = Sections.query.filter_by(id_version=version_id, section_number=3).order_by(desc(Sections.id)).all()

        unit_headers = {
            1: ('кг у.т.', 'т у.т.'),
            2: ('Мкал', 'Гкал'),
            3: ('кВтч', 'тыс.кВтч')
        }

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        def add_sheet(ws, sections, title, unit_header_one_text, unit_header_all_text):
            ws.title = title
            merged_cells = {
                'A1:A2': "Наименование вида продукции (работ услуг)",
                'B1:B2': "Код строки",
                'C1:C2': "Код по ОКЭД",
                'D1:D2': "Единица измерения",
                'E1:E2': "Произведено продукции (работ, услуг) за отчетный период",
                'F1:G1': f"Израсходовано на единицу продукции (работы, услуги) за отчетный период, {unit_header_one_text}",
                'H1:J1': f"Израсходовано на всю произведенную продукцию (работу, услугу) за отчетный период, {unit_header_all_text}",
                'K1:K2': "Примечание",
                'J1': "Экономия(-), перерасход(+)",
                'F2': "по утвержденной норме (предельному уровню)",
                'G2': "фактически",
                'H2': "по утвержденной норме (предельному уровню)",
                'I2': "фактически",
                'J2': "Экономия(-), перерасход(+)",
                'A3': "A",
                'B3': "Б",
                'C3': "В",
                'D3': "Г",
                'E3': "1",
                'F3': "2",
                'G3': "3",
                'H3': "4",
                'I3': "5",
                'J3': "6",
                'K3': "7"
            }

            font = Font(name='Times New Roman', size=12)
            alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            for cell_range, text in merged_cells.items():
                top_left_cell = cell_range.split(':')[0]
                ws[top_left_cell] = text
                ws[top_left_cell].font = font
                ws[top_left_cell].alignment = alignment
                ws[top_left_cell].border = border
            for cell_range in merged_cells.keys():
                ws.merge_cells(cell_range)
            column_widths = {
                'A': 30,
                'B': 15,
                'C': 15,
                'D': 20,
                'E': 50,
                'F': 40,
                'G': 40,
                'H': 40,
                'I': 40,
                'J': 30,
                'K': 20
            }
            row_heights = {
                1: 30,
                2: 30,
                3: 25
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            for row, height in row_heights.items():
                ws.row_dimensions[row].height = height

            row_index = 4
            for section in sections:
                ws[f'A{row_index}'] = section.product.NameProduct
                ws[f'B{row_index}'] = section.code_product
                ws[f'C{row_index}'] = section.Oked
                ws[f'D{row_index}'] = section.product.unit.NameUnit
                ws[f'E{row_index}'] = section.produced
                ws[f'F{row_index}'] = section.Consumed_Quota
                ws[f'G{row_index}'] = section.Consumed_Fact
                ws[f'H{row_index}'] = section.Consumed_Total_Quota
                ws[f'I{row_index}'] = section.Consumed_Total_Fact
                ws[f'J{row_index}'] = section.total_differents
                ws[f'K{row_index}'] = section.note

                for col in column_widths.keys():
                    cell = ws[f'{col}{row_index}']
                    cell.border = border
                row_index += 1

        unit_header_one_text1, unit_header_all_text1 = unit_headers.get(1, ('', ''))
        unit_header_one_text2, unit_header_all_text2 = unit_headers.get(2, ('', ''))
        unit_header_one_text3, unit_header_all_text3 = unit_headers.get(3, ('', ''))

        add_sheet(wb.create_sheet(), sections1, "Топливо", unit_header_one_text1, unit_header_all_text1)
        add_sheet(wb.create_sheet(), sections2, "Тепло", unit_header_one_text2, unit_header_all_text2)
        add_sheet(wb.create_sheet(), sections3, "Электричество", unit_header_one_text3, unit_header_all_text3)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name='table_report.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@auth.route('/print_ticket/<int:id>', methods=['POST'])
async def print_ticket(id):
    if request.method == 'POST':
        ticket = Ticket.query.get(id)
        version_report = ticket.version_report
        report = version_report.report
        if ticket is None:
            return flash("Квитанция не найдена","error")
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        
        font_path_bold = os.path.join(os.path.dirname(__file__), 'static', 'fonts', 'Montserrat-Bold.ttf')
        font_path_regular = os.path.join(os.path.dirname(__file__), 'static', 'fonts', 'Montserrat-Regular.ttf')
        pdfmetrics.registerFont(TTFont('MontserratBold', font_path_bold))
        pdfmetrics.registerFont(TTFont('MontserratRegular', font_path_regular))
        
        c.setFont("MontserratBold", 12)
        y_position = 730 
        
        data = {
            "Отчет по форме": "Сведения о нормах",
            "Период": f"Год: {report.year}; Квартал; {report.quarter}",
            "ОКПО предприятия": report.okpo,
            "Результат обработки": "Отчет принят в обработку" if ticket.luck else "Отчет не принят в обработку",
            "Дата обработки": ticket.begin_time.strftime("%Y-%m-%d"),
            "Причина отказа": " " if ticket.luck else ticket.note,
        }
        
        name_x = 100
        value_x = 250
        y_position -= 20
        
        for key, value in data.items():
            c.setFont("MontserratBold", 12)
            c.drawString(name_x, y_position, f"{key}:")
            c.setFont("MontserratRegular", 12)
            c.drawString(value_x, y_position, f"{value}")
            y_position -= 20
        c.save()
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name="ticket.pdf", mimetype="application/pdf")
    
@auth.route('/export_ready_reports', methods=['POST'])
async def export_ready_reports():
    if request.method == 'POST':
        year_filter = request.form.get('year_filter')
        quarter_filter = request.form.get('quarter_filter')

        filters = []
        if year_filter:
            filters.append(Report.year == year_filter)
        if quarter_filter:
            filters.append(Report.quarter == quarter_filter)

        # Проверка типа пользователя и ОКПО
        user_type = current_user.type
        user_organization_okpo = (
            str(current_user.organization.okpo)[-4] if current_user.organization else None
        )

        if user_type == "Администратор" or (
            current_user.organization and str(current_user.organization.okpo).startswith("8")
        ):
            # Администраторы или организации с ОКПО, начинающимся на 8, видят все
            versions = Version_report.query.options(
                joinedload(Version_report.report),
                joinedload(Version_report.sections).joinedload(Sections.product)
            ).join(Report).filter(
                Version_report.status == "Одобрен",
                *filters
            ).all()
        elif user_organization_okpo:
            # Фильтруем по четвёртой цифре ОКПО
            versions = Version_report.query.options(
                joinedload(Version_report.report),
                joinedload(Version_report.sections).joinedload(Sections.product)
            ).join(Report).filter(
                Version_report.status == "Одобрен",
                func.substr(func.cast(Report.okpo, String), -4, 1) == user_organization_okpo,
                *filters
            ).all()
        else:
            flash('У вас нет доступа к отчетам.', 'error')
            return redirect(url_for('views.audit_area', status='Одобрен', year=year_filter, quarter=quarter_filter))

        if not versions:
            flash('Отсутствуют одобренные отчеты для выбранных фильтров', 'error')
            return redirect(url_for('views.audit_area', status='Одобрен', year=year_filter, quarter=quarter_filter))

        # Генерация ZIP-архива с DBF-файлами
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED) as zip_file:
            for version in versions:
                report = version.report
                sections = version.sections
                data = [{
                    'INDX': str(section.id),
                    'YEAR_': str(report.year) if report.year is not None else '',
                    'KVARTAL': str(report.quarter) if report.quarter is not None else '',
                    'IDPREDPR': str(report.okpo) if report.okpo is not None else '',
                    'DATERECEIV': None,
                    'EXCEED': None,
                    'SECTIONNUM': str(section.section_number) if section.section_number is not None else '',
                    'CODEPROD': str(section.code_product) if section.code_product is not None else '',
                    'OKED': str(section.Oked) if section.Oked is not None else '',
                    'PRODUCED': str(section.produced) if section.produced is not None else '',
                    'CONSUMEDQ': str(section.Consumed_Quota) if section.Consumed_Quota is not None else '',
                    'CONSUMEDF': str(section.Consumed_Fact) if section.Consumed_Fact is not None else '',
                    'CONSUMEDQT': str(section.Consumed_Total_Quota) if section.Consumed_Total_Quota is not None else '',
                    'CONSUMEDFT': str(section.Consumed_Total_Fact) if section.Consumed_Total_Fact is not None else '',
                    'COMMENT1': section.note if section.note is not None else '',
                    'COMMENT2': None,
                    'NAMEPROD': section.product.NameProduct[:200].encode('cp866', 'replace').decode('cp866') if section.product and section.product.NameProduct else '',
                    'CODEUNIT': None,
                    'NAMEORG': report.organization_name[:200].encode('cp866', 'replace').decode('cp866') if report.organization_name is not None else '',
                } for section in sections]

                df = pd.DataFrame(data)

                with NamedTemporaryFile(delete=False, suffix='.dbf') as temp_file:
                    temp_filename = temp_file.name

                    table = dbf.Table(
                        temp_filename,
                        'INDX C(10); YEAR_ C(20); KVARTAL C(20); IDPREDPR C(20); DATERECEIV C(20); '
                        'EXCEED C(20); SECTIONNUM C(20); CODEPROD C(20); OKED C(20); PRODUCED C(20); '
                        'CONSUMEDQ C(20); CONSUMEDF C(20); CONSUMEDQT C(20); CONSUMEDFT C(20); '
                        'COMMENT1 C(20); COMMENT2 C(20); NAMEPROD C(200); CODEUNIT C(20); NAMEORG C(200); ',
                        codepage='cp866'
                    )

                    table.open(mode=dbf.READ_WRITE)

                    try:
                        for _, row in df.iterrows():
                            row_dict = row.to_dict()
                            table.append(row_dict)
                    finally:
                        table.close()

                    with open(temp_filename, 'rb') as f:
                        dbf_filename = f'{report.okpo}_{report.year}_{report.quarter}.dbf'
                        zip_file.writestr(dbf_filename, f.read())

        zip_buffer.seek(0)
        return Response(
            zip_buffer,
            mimetype='application/zip',
            headers={"Content-Disposition": "attachment;filename=reports.zip"}
        )

@auth.route('/sent_for_admin', methods=['POST'])
async def sent_for_admin():
    if request.method == 'POST':
        text = request.form.get('text')
        if text:
            admins = User.query.filter_by(type = "Администратор").all()
            if admins:
                for i in admins:
                    new_message = Message(
                        sender = current_user.id,
                        text = text,
                        user_id = i.id
                    )
                    db.session.add(new_message)
                    db.session.commit()
                flash('Сообщение отправлено', 'succes')
            else:
                flash('Администраторов нет', 'error')
        else:
            flash('Пустое сообщение', 'error')
    return redirect(url_for('views.beginPage'))